# excel_handler.py
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from link_parser import parse_hyperlink, extract_filename_from_link
from models import FileInfo

def load_excel_sheets(filepath: str):
    """Lấy danh sách sheet names"""
    wb = load_workbook(filepath, data_only=False)
    return wb.sheetnames


def get_sheet_data(filepath: str, sheet_name: str, 
                   title_col: str, data_col: str,
                   start_row: int = 1, max_rows: int = 100):
    """
    Lấy dữ liệu từ sheet để preview
    Trả về: (headers, rows) với rows là list of dict
    """
    wb = load_workbook(filepath, data_only=False)
    ws = wb[sheet_name]
    
    # Xác định cột index
    title_col_idx = ord(title_col.upper()) - ord('A') + 1 if title_col else 1
    data_col_idx = ord(data_col.upper()) - ord('A') + 1 if data_col else 3
    
    # Lấy tiêu đề cột (dòng 1)
    headers = []
    for col in range(1, min(ws.max_column + 1, 10)):  # Giới hạn 10 cột preview
        cell_value = ws.cell(row=1, column=col).value
        headers.append(str(cell_value) if cell_value else chr(64 + col))
    
    # Lấy dữ liệu các dòng
    rows = []
    end_row = min(start_row + max_rows - 1, ws.max_row)
    
    for row in range(start_row, end_row + 1):
        row_data = {}
        for col in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row=row, column=col)
            cell_value = cell.value
            
            # Nếu là cột dữ liệu, thử parse link
            if col == data_col_idx:
                link = parse_hyperlink(cell)
                row_data[chr(64 + col)] = link if link else (cell_value if cell_value else "")
                
                # Lấy tiêu đề từ cột title
                title_cell = ws.cell(row=row, column=title_col_idx)
                row_data["title"] = title_cell.value if title_cell.value else f"Dòng {row}"
            else:
                row_data[chr(64 + col)] = str(cell_value)[:50] if cell_value else ""
        
        rows.append(row_data)
    
    wb.close()
    return headers, rows, ws.max_row


def get_links_from_sheet(filepath: str, sheet_name: str, 
                         title_col: str, data_col: str,
                         full_sheet: bool = True, current_row: int = 1) -> list[FileInfo]:
    """
    Lấy danh sách link từ sheet
    Nếu full_sheet=False chỉ lấy từ current_row
    """
    wb = load_workbook(filepath, data_only=False)
    ws = wb[sheet_name]
    
    title_col_idx = ord(title_col.upper()) - ord('A') + 1
    data_col_idx = ord(data_col.upper()) - ord('A') + 1
    
    files = []
    start_row = 2 if full_sheet else current_row  # Dòng 1 là tiêu đề
    end_row = ws.max_row if full_sheet else current_row
    
    for row in range(start_row, end_row + 1):
        data_cell = ws.cell(row=row, column=data_col_idx)
        title_cell = ws.cell(row=row, column=title_col_idx)
        
        link = parse_hyperlink(data_cell)
        title = title_cell.value if title_cell.value else f"Dòng {row}"
        
        if link:
            files.append(FileInfo(
                sheet=sheet_name,
                row=row,
                title=str(title),
                old_link=link
            ))
    
    wb.close()
    return files


def update_excel_link(filepath: str, sheet_name: str, row: int, 
                      data_col: str, new_link: str):
    """Cập nhật link mới vào ô Excel"""
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    
    data_col_idx = ord(data_col.upper()) - ord('A') + 1
    cell = ws.cell(row=row, column=data_col_idx)
    cell.value = new_link
    cell.hyperlink = new_link
    
    wb.save(filepath)
    wb.close()


def highlight_column(filepath: str, sheet_name: str, column_letter: str):
    """Tô màu cột trong preview (chỉ để hiển thị, không lưu)"""
    # Chức năng này chỉ để UI, không cần implement ở đây
    pass