# excel/pyxlsb_reader.py
import os
from pyxlsb import open_workbook
from .reader import ExcelReader

class PyxlsbReader(ExcelReader):
    """Hỗ trợ .xlsb - đọc được cả công thức"""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = open_workbook(filepath)
        self._sheet_cache = {}
        self._formula_cache = {}
        self._modified = False
        self._changes = []
    
    def get_sheet_names(self) -> list[str]:
        try:
            return self.wb.sheets
        except Exception as e:
            print(f"Warning: get_sheet_names failed: {e}")
            return []
    
    def _get_sheet_data(self, sheet_name: str):
        """Lấy toàn bộ dữ liệu sheet vào cache"""
        if sheet_name not in self._sheet_cache:
            sheet_data = []
            try:
                with self.wb.get_sheet(sheet_name) as sheet:
                    for row_idx, row in enumerate(sheet.rows(), start=1):
                        row_data = []
                        for col_idx, cell in enumerate(row, start=1):
                            row_data.append(cell)
                        sheet_data.append(row_data)
                self._sheet_cache[sheet_name] = sheet_data
            except Exception as e:
                print(f"Warning: _get_sheet_data failed ({sheet_name}): {e}")
                self._sheet_cache[sheet_name] = []
        return self._sheet_cache[sheet_name]
    
    def get_cell_value(self, sheet_name: str, row: int, col: int):
        try:
            sheet_data = self._get_sheet_data(sheet_name)
            if row - 1 < len(sheet_data) and col - 1 < len(sheet_data[row - 1]):
                cell = sheet_data[row - 1][col - 1]
                # pyxlsb trả về object có attribute 'v' (value)
                if hasattr(cell, 'v'):
                    return cell.v
                return cell
            return None
        except Exception as e:
            print(f"Warning: get_cell_value failed ({sheet_name},{row},{col}): {e}")
            return None
    
    def get_cell_formula(self, sheet_name: str, row: int, col: int) -> str | None:
        try:
            sheet_data = self._get_sheet_data(sheet_name)
            if row - 1 < len(sheet_data) and col - 1 < len(sheet_data[row - 1]):
                cell = sheet_data[row - 1][col - 1]
                # Thử nhiều cách để lấy công thức
                if hasattr(cell, 'formula') and cell.formula:
                    return str(cell.formula)
                if hasattr(cell, 'f') and cell.f:
                    return str(cell.f)
                # Nếu không, thử lấy từ value (có thể là công thức)
                if hasattr(cell, 'v') and cell.v:
                    return str(cell.v)
            return None
        except Exception as e:
            print(f"Warning: get_cell_formula failed ({sheet_name},{row},{col}): {e}")
            return None
    
    def has_formula(self, sheet_name: str, row: int, col: int) -> bool:
        try:
            formula = self.get_cell_formula(sheet_name, row, col)
            return formula is not None
        except:
            return False
    
    def get_hyperlink(self, sheet_name: str, row: int, col: int) -> str | None:
        try:
            # Thử lấy từ giá trị nếu là text thuần
            value = self.get_cell_value(sheet_name, row, col)
            if isinstance(value, str) and (value.startswith('\\\\') or ':\\' in value or value.startswith('http')):
                return value
            return None
        except Exception as e:
            print(f"Warning: get_hyperlink failed ({sheet_name},{row},{col}): {e}")
            return None
    
    def get_max_row(self, sheet_name: str) -> int:
        try:
            sheet_data = self._get_sheet_data(sheet_name)
            return len(sheet_data)
        except:
            return 0
    
    def get_max_col(self, sheet_name: str) -> int:
        try:
            sheet_data = self._get_sheet_data(sheet_name)
            if sheet_data:
                return len(sheet_data[0])
            return 0
        except:
            return 0
    
    def update_cell(self, sheet_name: str, row: int, col: int, value: str):
        """Lưu thay đổi để ghi sau"""
        self._changes.append({
            'sheet': sheet_name,
            'row': row,
            'col': col,
            'value': value
        })
        self._modified = True
    
    def save(self, filepath: str = None):
        if not self._modified:
            return
        
        if filepath is None:
            base = os.path.splitext(self.filepath)[0]
            filepath = f"{base}_updated.xlsx"
        
        try:
            from openpyxl import Workbook
            wb_new = Workbook()
            
            for sheet_name in self.get_sheet_names():
                if sheet_name == wb_new.sheetnames[0] and len(wb_new.sheetnames) == 1:
                    ws_new = wb_new.active
                    ws_new.title = sheet_name
                else:
                    ws_new = wb_new.create_sheet(sheet_name)
                
                max_row = self.get_max_row(sheet_name)
                max_col = self.get_max_col(sheet_name)
                
                for r in range(1, max_row + 1):
                    for c in range(1, max_col + 1):
                        value = self.get_cell_value(sheet_name, r, c)
                        if value is not None:
                            ws_new.cell(row=r, column=c, value=value)
                
                for change in self._changes:
                    if change['sheet'] == sheet_name:
                        ws_new.cell(row=change['row'], column=change['col'], value=change['value'])
            
            wb_new.save(filepath)
            print(f"Đã lưu file mới: {filepath}")
        except Exception as e:
            print(f"Warning: save failed: {e}")
    
    def close(self):
        try:
            self.wb.close()
        except:
            pass