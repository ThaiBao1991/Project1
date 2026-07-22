# excel/openpyxl_reader.py
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from .reader import ExcelReader

class OpenpyxlReader(ExcelReader):
    """Hỗ trợ .xlsx, .xlsm, .xltx, .xltm"""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = load_workbook(filepath, data_only=False)
    
    def get_sheet_names(self) -> list[str]:
        return self.wb.sheetnames
    
    def _get_worksheet(self, sheet_name: str):
        return self.wb[sheet_name]
    
    def get_cell_value(self, sheet_name: str, row: int, col: int):
        try:
            ws = self._get_worksheet(sheet_name)
            cell = ws.cell(row=row, column=col)
            return cell.value
        except Exception as e:
            print(f"Warning: get_cell_value failed ({sheet_name},{row},{col}): {e}")
            return None
    
    def get_cell_formula(self, sheet_name: str, row: int, col: int) -> str | None:
        try:
            ws = self._get_worksheet(sheet_name)
            cell = ws.cell(row=row, column=col)
            
            # Cách đúng để kiểm tra công thức trong openpyxl
            if cell.data_type == 'f':
                # Trong openpyxl, công thức được lưu trong cell.value
                return str(cell.value) if cell.value else None
            return None
        except Exception as e:
            print(f"Warning: get_cell_formula failed ({sheet_name},{row},{col}): {e}")
            return None
    
    def has_formula(self, sheet_name: str, row: int, col: int) -> bool:
        try:
            ws = self._get_worksheet(sheet_name)
            cell = ws.cell(row=row, column=col)
            return cell.data_type == 'f'
        except Exception as e:
            print(f"Warning: has_formula failed ({sheet_name},{row},{col}): {e}")
            return False
    
    def get_hyperlink(self, sheet_name: str, row: int, col: int) -> str | None:
        try:
            ws = self._get_worksheet(sheet_name)
            cell = ws.cell(row=row, column=col)
            if cell.hyperlink and cell.hyperlink.target:
                return cell.hyperlink.target
            return None
        except Exception as e:
            print(f"Warning: get_hyperlink failed ({sheet_name},{row},{col}): {e}")
            return None
    
    def get_max_row(self, sheet_name: str) -> int:
        try:
            ws = self._get_worksheet(sheet_name)
            return ws.max_row
        except Exception as e:
            print(f"Warning: get_max_row failed ({sheet_name}): {e}")
            return 0
    
    def get_max_col(self, sheet_name: str) -> int:
        try:
            ws = self._get_worksheet(sheet_name)
            return ws.max_column
        except Exception as e:
            print(f"Warning: get_max_col failed ({sheet_name}): {e}")
            return 0
    
    def update_cell(self, sheet_name: str, row: int, col: int, value: str):
        """Cập nhật giá trị ô"""
        try:
            ws = self._get_worksheet(sheet_name)
            cell = ws.cell(row=row, column=col)
            cell.value = value
            # Nếu là đường dẫn, tạo hyperlink
            if value and (value.startswith('\\\\') or ':\\' in value or value.startswith('http')):
                cell.hyperlink = value
            self._modified = True
        except Exception as e:
            print(f"Warning: update_cell failed ({sheet_name},{row},{col}): {e}")
            raise
    
    def save(self, filepath: str = None):
        try:
            if filepath is None:
                filepath = self.filepath
            self.wb.save(filepath)
        except Exception as e:
            print(f"Warning: save failed: {e}")
    
    def close(self):
        try:
            self.wb.close()
        except:
            pass